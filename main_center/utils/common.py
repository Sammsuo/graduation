import os
import xlrd
import openpyxl
import re
import json
from datetime import datetime
from xml.etree import ElementTree as ElementTree
import datetime
from main_center.utils import configDB

proDir = os.path.split(os.path.realpath(__file__))[0]
localDB = configDB.MyDB()


def check_url(url):
    check_rule_r = r'(https?:.*?:.*?/)'
    check_rule_p = re.compile(check_rule_r)
    check_rule_c = re.findall(check_rule_p, url)
    print(check_rule_c)
    if len(check_rule_c) != 0:
        return True
    else:
        return False


def show_return_msg(response):
    msg = response.text
    print("\n请求返回值: " + '\n' + json.dumps(json.loads(msg), ensure_ascii=False, sort_keys=True, indent=4))


# def get_report_path():
#     """
#     获取报告文件路径
#     :return:
#     """
#     report_path = os.path.join(reportPath, 'report.html')
#     return report_path
#
#
# def get_result_path():
#     """
#     获取报告结果路径
#     :return:
#     """
#     return reportPath

def get_file():
    """
    获取excel文件
    :return:
    """
    path = os.path.join(proDir, 'testFile')
    dirs = os.listdir(path)
    f = []
    for i in dirs:
        if i.endswith(".xlsx"):
            f.append(i)
    return f[0]


def get_xlsx_sheets(xlsx_name):
    cls = []
    # get xls file's path
    print('llalala:', xlsx_name)
    xls_path = os.path.join(proDir, "testFile", xlsx_name)
    # print (xls_path)

    # open xls file
    file = xlrd.open_workbook(xls_path)

    sheets = file.sheets()

    for sheet in sheets:
        nrows = sheet.nrows
        for i in range(nrows):
            if sheet.row_values(i)[0] != u'case_name':
                cls.append(sheet.row_values(i))
    return cls


def delete_file():
    """
    删除文件
    :return:
    """
    delete_path = os.path.join(proDir, 'testFile')
    dirs = os.listdir(delete_path)
    for i in dirs:
        if i.endswith('.xlsx') or i.endswith('.xls'):
            os.remove(os.path.join(delete_path, i))
    return '删除完毕'


def delete_file_by_name(name, path):
    """
    通过文件名删除文件
    :param name:
    :return:
    """
    delete_path = os.path.join(proDir, path)
    dirs = os.listdir(delete_path)
    for i in dirs:
        if i == name:
            os.remove(os.path.join(delete_path, i))
            return True
        else:
            return False

# ----------------------------------------------------------


def download_file(lists, flag):
    """
    下载excel文件
    :return:
    """
    # print(req.data) # 获取data属性(传一个列表)
    get_list = lists
    download_path = os.path.join(proDir, 'download_template', choose_temp(flag))
    if flag == '1':
        finish_path = os.path.join(proDir, 'download_template', 'Func_case.xlsx')
    elif flag == '2':
        finish_path = os.path.join(proDir, 'download_template', 'API_case.xlsx')
    # print(download_path)
    wb = openpyxl.load_workbook(download_path)
    table = wb['Sheet1']
    # 顺序填写，序号，名字，前置，步骤，预期，备注
    row = 2
    column = 2
    num = 1
    for case_list in get_list:
        if case_list is not None:
            for case_parm in case_list:
                table.cell(row=row, column=1).value = num
                table.cell(row=row, column=column).value = case_parm
                column += 1
            row += 1
            column = 2
            num += 1
    wb.save(finish_path)
    return finish_path

def choose_temp(flag):
    """
    选择模板
    :param flag:
    :return:
    """
    if flag == '1':
        return 'func_template_01.xlsx'
    elif flag == '2':
        return 'api_template_02.xlsx'


# ########################### 读数据库xml ##############################
database = {}


def set_xml():
    """
    set sql xml
    :return:
    """
    if len(database) == 0:
        sql_path = os.path.join(proDir, "testFile", "SQL.xml")
        tree = ElementTree.parse(sql_path)
        for db in tree.findall("database"):
            db_name = db.get("name")
            table = {}
            for tb in db.getchildren():
                table_name = tb.get("name")
                sql = {}
                for data in tb.getchildren():
                    sql_id = data.get("id")
                    sql[sql_id] = data.text
                table[table_name] = sql
            database[db_name] = table


def get_xml_dict(database_name, table_name):
    """
    get db dict by given name
    :param database_name:
    :param table_name:
    :return:
    """
    set_xml()
    database_dict = database.get(database_name).get(table_name)
    return database_dict


def get_sql(database_name, table_name, sql_id):
    """
    get sql by given name and sql_id
    :param database_name:
    :param table_name:
    :param sql_id:
    :return:
    """

    db = get_xml_dict(database_name, table_name)
    sql = db.get(sql_id)
    return sql

# ##################################################################
def changge_check_params(str):
    """
    change excel's check_param (str) to be list, return dic
    :param str:
    :return:
    """
    key_value_list = str.split(';')
    # print(key_value_list)
    dic = {}
    for i in key_value_list:
        change_list = i.split('=')
        change_list = list(map(lambda x: x.strip(), change_list))
        # print(change_list)
        dic[change_list[0]] = change_list[1]
    return dic


def change_params_style(params_list, module):
    try:
        if module == '4':
            res_list = []
            for i in params_list:
                set_dict = {}
                set_dict['case_title'] = i['CaseName']
                set_dict['case_condition'] = '前置条件: ' + i['CaseCondition'] + '  备注: ' + i['CaseRemark']
                set_dict['module'] = module
                set_dict['case_step'] = i['CaseStep']
                set_dict['case_expect'] = i['CaseExpect']
                print(set_dict)
                res_list.append(set_dict)
            return res_list
        elif module == '3':
            res_list = []
            for i in params_list:
                set_dict = {}
                set_dict['case_title'] = i['CaseName']
                set_dict['case_condition'] = 'URL: ' + i['CaseUrl'] + '  Method: ' + i['CaseMethod'] + ' Header: ' + i['CaseHeader']
                set_dict['module'] = module
                set_dict['case_step'] = i['CaseParams']
                set_dict['case_expect'] = i['CaseResult']
                print(set_dict)
                res_list.append(set_dict)
            return res_list
    except Exception as e:
        return e


def execute_upload(upload_list):
    print(upload_list)
    print(type(upload_list))
    if type(upload_list) == list:
        for case in upload_list:
            insert_case_params = [case['module'], case['case_title'], case['case_condition'], datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            localDB.insert_zt_case(*insert_case_params)
            case_id = localDB.select_id_by_openedDate()
            insert_case_step_param = [case_id, case['case_step'], case['case_expect']]
            localDB.insert_zt_casestep(*insert_case_step_param)


def get_line_data():
    month_data = {'Jan': 0, 'Feb': 0, 'Mar': 0, 'Apr': 0, 'May': 0, 'June': 0, 'July': 0, 'Aug': 0, 'Sept': 0, 'Oct': 0, 'Nov': 0, 'Dec': 0}
    month_tuple = localDB.conut_bug_by_month()
    x = 0
    for i in month_data:
        if x < len(month_tuple):
            month_data[i] = month_tuple[x][1]
            x += 1
    return month_data


def get_bug_style():
    style_dict = {'codeerror': 0, 'designdefect': 0, 'others': 0}
    style_get_dict = localDB.get_bug_style()
    for i in style_get_dict:
        style_dict[i[0]] = i[1]

    return style_dict


def get_bug_module():
    module_dict = {'all': 0, 'up': 0, 'down': 0}
    all = localDB.get_bug_all()
    up = localDB.get_bug_up()
    down = localDB.get_bug_down()
    localDB.closeDB()
    module_dict['all'] = all[0]
    module_dict['up'] = up[0]
    module_dict['down'] = down[0]
    print(all, up, down)
    return module_dict


if __name__ == '__main__':
    # print(check_url('http://47.107.21.127:9000/pld/credit/#/credit/packageManage/index'))
    # c = changge_check_params("code = '00000000';msg='成功'")
    # print(type(c['code']))
    # print(type(c['msg']))
    # print(proDir)
    get_line_data()
